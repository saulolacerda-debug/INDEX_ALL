from __future__ import annotations

from collections import Counter
from datetime import date, datetime, time, timedelta
from pathlib import Path

from openpyxl import load_workbook

from index_all.parsers.legal_structure import (
    StructuredTextRecord,
    build_legal_blocks,
    looks_like_legal_document,
    make_preview_title,
    normalize_text,
)


def _json_safe_value(value: object) -> object:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, timedelta):
        return str(value)
    return str(value)


def _render_values(values: list[object]) -> str:
    rendered_values = [str(value) for value in values if value is not None and str(value).strip()]
    if not rendered_values:
        return ""
    return normalize_text(" | ".join(rendered_values))


def parse_xlsx(path: Path) -> dict:
    workbook = load_workbook(filename=str(path), data_only=True, read_only=True)
    try:
        sheet_payloads: list[dict[str, object]] = []
        records: list[StructuredTextRecord] = []
        for sheet in workbook.worksheets:
            sheet_rows: list[dict[str, object]] = []
            row_count = 0
            max_column = 0

            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                row_count = row_number
                values = list(row)
                max_column = max(max_column, len(values))
                safe_values = [_json_safe_value(value) for value in values]
                rendered = _render_values(safe_values)
                if not rendered:
                    continue
                sheet_rows.append(
                    {
                        "row_number": row_number,
                        "safe_values": safe_values,
                        "rendered": rendered,
                    }
                )
                records.append(
                    StructuredTextRecord(
                        text=rendered,
                        locator={"page": None, "sheet": sheet.title, "line_start": row_number, "line_end": row_number},
                        extra={"sheet": sheet.title, "values": safe_values},
                    )
                )
            sheet_payloads.append(
                {
                    "sheet_name": sheet.title,
                    "row_count": row_count,
                    "column_count": max_column,
                    "rows": sheet_rows,
                }
            )

        if looks_like_legal_document([record.text for record in records]):
            blocks = build_legal_blocks(records)
            mode = "structured_legal"
        else:
            blocks = []
            block_idx = 1
            total_row_count = 0
            total_non_empty_row_count = 0
            sheet_stats: list[dict[str, int | str]] = []

            for sheet_payload in sheet_payloads:
                sheet_name = str(sheet_payload["sheet_name"])
                row_count = int(sheet_payload["row_count"])
                column_count = int(sheet_payload["column_count"])
                sheet_rows = list(sheet_payload["rows"])
                non_empty_row_count = len(sheet_rows)
                total_row_count += row_count

                blocks.append(
                    {
                        "id": f"block_{block_idx:04d}",
                        "kind": "sheet",
                        "title": sheet_name,
                        "text": f"Sheet {sheet_name}",
                        "locator": {"page": None, "sheet": sheet_name, "line_start": None, "line_end": None},
                        "extra": {"max_row": row_count, "max_column": column_count},
                    }
                )
                block_idx += 1

                for row_payload in sheet_rows:
                    row_number = int(row_payload["row_number"])
                    safe_values = list(row_payload["safe_values"])
                    rendered = str(row_payload["rendered"])
                    blocks.append(
                        {
                            "id": f"block_{block_idx:04d}",
                            "kind": "sheet_row",
                            "title": make_preview_title(rendered),
                            "text": rendered,
                            "locator": {"page": None, "sheet": sheet_name, "line_start": row_number, "line_end": row_number},
                            "extra": {"values": safe_values},
                        }
                    )
                    block_idx += 1

                total_non_empty_row_count += non_empty_row_count
                sheet_stats.append(
                    {
                        "sheet_name": sheet_name,
                        "row_count": row_count,
                        "non_empty_row_count": non_empty_row_count,
                        "column_count": column_count,
                    }
                )
            mode = "sheet_full"

        parser_metadata = {
            "sheet_names": workbook.sheetnames,
            "sheet_count": len(workbook.worksheets),
            "mode": mode,
            "block_count": len(blocks),
            "kind_counts": dict(sorted(Counter(block["kind"] for block in blocks).items())),
        }
        if mode == "sheet_full":
            parser_metadata["row_count"] = total_row_count
            parser_metadata["non_empty_row_count"] = total_non_empty_row_count
            parser_metadata["sheet_stats"] = sheet_stats

        return {
            "content": {
                "blocks": blocks,
                "parser_metadata": parser_metadata,
            }
        }
    finally:
        workbook.close()
