from __future__ import annotations

import csv
import json
from datetime import datetime

from openpyxl import Workbook

from index_all.parsers.csv_parser import parse_csv
from index_all.parsers.xlsx_parser import parse_xlsx

from tests.helpers import workspace_test_dir


def test_csv_parser_indexes_all_rows_for_tabular_documents():
    with workspace_test_dir() as temp_dir:
        path = temp_dir / "amostra.csv"
        with path.open("w", encoding="utf-8", newline="") as file_obj:
            writer = csv.writer(file_obj)
            writer.writerow(["conta", "valor", "status"])
            for index in range(1, 26):
                writer.writerow([f"conta_{index}", index * 10, "ativo" if index % 2 else "pendente"])

        result = parse_csv(path)
        blocks = result["content"]["blocks"]
        parser_metadata = result["content"]["parser_metadata"]

        assert parser_metadata["mode"] == "table_full"
        assert parser_metadata["row_count"] == 26
        assert parser_metadata["non_empty_row_count"] == 26
        assert parser_metadata["data_row_count"] == 25
        assert parser_metadata["block_count"] == 26
        assert blocks[0]["kind"] == "table_header"
        assert len([block for block in blocks if block["kind"] == "table_row"]) == 25
        assert blocks[-1]["locator"]["line_start"] == 26
        assert "conta_25" in blocks[-1]["text"]


def test_xlsx_parser_indexes_all_rows_for_tabular_documents():
    with workspace_test_dir() as temp_dir:
        path = temp_dir / "amostra.xlsx"
        workbook = Workbook()
        primeira_aba = workbook.active
        primeira_aba.title = "Financeiro"
        for row_number in range(1, 13):
            primeira_aba.cell(row=row_number, column=1).value = f"financeiro_{row_number}"
            if row_number == 3:
                primeira_aba.cell(row=row_number, column=2).value = datetime(2026, 3, 15, 10, 30, 0)
            else:
                primeira_aba.cell(row=row_number, column=2).value = row_number * 100

        segunda_aba = workbook.create_sheet("Socios")
        for row_number in range(1, 5):
            segunda_aba.cell(row=row_number, column=1).value = f"socio_{row_number}"
            segunda_aba.cell(row=row_number, column=2).value = "ativo"
        workbook.save(path)
        workbook.close()

        result = parse_xlsx(path)
        blocks = result["content"]["blocks"]
        parser_metadata = result["content"]["parser_metadata"]

        assert parser_metadata["mode"] == "sheet_full"
        assert parser_metadata["sheet_count"] == 2
        assert parser_metadata["row_count"] == 16
        assert parser_metadata["non_empty_row_count"] == 16
        assert parser_metadata["block_count"] == 18
        assert [item["sheet_name"] for item in parser_metadata["sheet_stats"]] == ["Financeiro", "Socios"]
        assert parser_metadata["sheet_stats"][0]["non_empty_row_count"] == 12
        assert parser_metadata["sheet_stats"][1]["non_empty_row_count"] == 4
        assert len([block for block in blocks if block["kind"] == "sheet"]) == 2
        assert len([block for block in blocks if block["kind"] == "sheet_row"]) == 16
        assert blocks[3]["extra"]["values"][1] == "2026-03-15T10:30:00"
        assert any(block["locator"]["sheet"] == "Socios" and block["locator"]["line_start"] == 4 for block in blocks)
        json.dumps(result, ensure_ascii=False)
